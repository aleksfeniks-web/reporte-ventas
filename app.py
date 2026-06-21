from flask import Flask, request, send_file, Response, jsonify
import json, io, datetime, os
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ── Database ──────────────────────────────────────────────────────────────────

def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'], sslmode='require')

def init_db():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS entradas (
                    id SERIAL PRIMARY KEY,
                    fecha TEXT NOT NULL,
                    nombre TEXT NOT NULL,
                    razon_social TEXT,
                    nombre_obra TEXT,
                    telefono TEXT,
                    ubicacion TEXT,
                    tipos JSONB DEFAULT '[]',
                    motivos JSONB DEFAULT '[]',
                    otro_detalle TEXT,
                    m3_concreto TEXT,
                    no_venta_block TEXT,
                    creado_en TIMESTAMP DEFAULT NOW()
                )
            """)
            cur.execute("ALTER TABLE entradas ADD COLUMN IF NOT EXISTS importe NUMERIC DEFAULT 0;")
            cur.execute("ALTER TABLE entradas ADD COLUMN IF NOT EXISTS descuento NUMERIC DEFAULT 0;")
            cur.execute("ALTER TABLE entradas ADD COLUMN IF NOT EXISTS comision NUMERIC DEFAULT 0;")
        conn.commit()

with app.app_context():
    try:
        init_db()
    except Exception as e:
        print(f"DB init error: {e}")

# ── Excel generation ──────────────────────────────────────────────────────────

GRAY_FILL  = PatternFill("solid", fgColor="D9D9D9")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

def _s(style='thin'): return Side(style=style)

def _label(ws, cell_ref, text, underline=False):
    c = ws[cell_ref]
    c.value = text
    c.font  = Font(bold=True, size=9, name='Arial',
                   underline='single' if underline else None)
    c.fill  = GRAY_FILL
    c.alignment = Alignment(wrap_text=True, vertical='center')

def _value(ws, cell_ref, text, center=False, size=9):
    c = ws[cell_ref]
    c.value = text
    c.font  = Font(size=size, name='Arial')
    c.fill  = WHITE_FILL
    c.alignment = Alignment(
        wrap_text=True, vertical='center',
        horizontal='center' if center else 'left'
    )

def _x(ws, cell_ref, mark):
    c = ws[cell_ref]
    c.value = mark
    c.font  = Font(bold=True, size=11, name='Arial')
    c.fill  = WHITE_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')

def write_entry(ws, r, e):
    tipos   = e.get('tipos', [])
    motivos = e.get('motivos', [])

    ws.row_dimensions[r  ].height = 36
    ws.row_dimensions[r+1].height = 24
    ws.row_dimensions[r+2].height = 14
    ws.row_dimensions[r+3].height = 18
    ws.row_dimensions[r+4].height = 14
    ws.row_dimensions[r+5].height = 14

    _label(ws, f'A{r}', 'Fecha')
    ws.merge_cells(f'B{r}:D{r}')
    _value(ws, f'B{r}', e.get('fecha', ''))
    _label(ws, f'E{r}', 'Obra')
    _x(ws,    f'F{r}', 'X' if 'Obra' in tipos else '')
    _label(ws, f'G{r}', 'Referencia')
    _x(ws,    f'H{r}', 'X' if 'Referencia' in motivos else '')
    ws.merge_cells(f'I{r}:J{r}')
    _label(ws, f'I{r}', 'M³ venta\nConcreto o\nPotenciales')
    ws.merge_cells(f'K{r}:K{r+1}')
    _value(ws, f'K{r}', e.get('m3Concreto', ''), center=True, size=11)

    _label(ws, f'A{r+1}', 'Nombre:\nCliente o Prospecto')
    ws.merge_cells(f'B{r+1}:D{r+1}')
    _value(ws, f'B{r+1}', e.get('nombre', ''))
    _label(ws, f'E{r+1}', 'Auto\nconstruccion')
    _x(ws,    f'F{r+1}', 'X' if 'Autoconstruccion' in tipos else '')
    _label(ws, f'G{r+1}', 'Visita')
    _x(ws,    f'H{r+1}', 'X' if 'Visita' in motivos else '')

    _label(ws, f'A{r+2}', 'Razon Social')
    ws.merge_cells(f'B{r+2}:D{r+2}')
    _value(ws, f'B{r+2}', e.get('razonSocial', ''))
    _label(ws, f'E{r+2}', 'Reventa')
    _x(ws,    f'F{r+2}', 'X' if 'Reventa' in tipos else '')
    _label(ws, f'G{r+2}', 'Red Social')
    _x(ws,    f'H{r+2}', 'X' if 'Red Social' in motivos else '')
    ws.merge_cells(f'I{r+2}:J{r+2}')
    _label(ws, f'I{r+2}', 'No venta\nBlock o\nPotenciales')
    ws.merge_cells(f'K{r+2}:K{r+3}')
    _value(ws, f'K{r+2}', e.get('noVentaBlock', ''), center=True, size=11)

    _label(ws, f'A{r+3}', 'Nombre\nde la Obra')
    ws.merge_cells(f'B{r+3}:D{r+3}')
    _value(ws, f'B{r+3}', e.get('nombreObra', ''))
    _label(ws, f'E{r+3}', 'Evento')
    _x(ws,    f'F{r+3}', 'X' if 'Evento' in tipos else '')
    otro = e.get('otroDetalle', '')
    _label(ws, f'G{r+3}', f'Otro: {otro}' if otro else 'Otro:', underline=True)
    _x(ws,    f'H{r+3}', 'X' if 'Otro' in motivos else '')

    _label(ws, f'A{r+4}', 'Telefono')
    ws.merge_cells(f'B{r+4}:D{r+4}')
    _value(ws, f'B{r+4}', e.get('telefono', ''))

    # Sales metrics
    importe = float(e.get('importe') or 0)
    descuento = float(e.get('descuento') or 0)
    comision = float(e.get('comision') or 0)
    venta_neta = importe * (1 - descuento / 100.0)

    _label(ws, f'E{r+4}', 'Importe')
    _value(ws, f'G{r+4}', f"${importe:,.2f}")
    ws.merge_cells(f'I{r+4}:J{r+4}')
    _label(ws, f'I{r+4}', 'Descuento')
    _value(ws, f'K{r+4}', f"{descuento:.1f}%", center=True)

    _label(ws, f'A{r+5}', 'Ubicación')
    ws.merge_cells(f'B{r+5}:D{r+5}')
    _value(ws, f'B{r+5}', e.get('ubicacion', ''))

    _label(ws, f'E{r+5}', 'Comisión')
    _value(ws, f'G{r+5}', f"${comision:,.2f}")
    ws.merge_cells(f'I{r+5}:J{r+5}')
    _label(ws, f'I{r+5}', 'Venta Neta')
    _value(ws, f'K{r+5}', f"${venta_neta:,.2f}", center=True)

    thin, dashed, COLS = 'thin', 'dashed', 11
    for ri in range(r, r+6):
        for ci in range(1, COLS+1):
            cell = ws.cell(row=ri, column=ci)
            top    = thin if ri == r    else None
            bottom = thin if ri == r+5 else dashed
            left   = thin if ci == 1   else None
            right  = thin if ci == COLS else None
            if ci == 4:  right = thin
            if ci == 5:  left  = thin
            if ci == 6:  right = thin
            if ci == 7:  left  = thin
            if ci == 8:  right = thin
            if ci == 9:  left  = thin
            if ci == 10: right = thin
            if ci == 11: left  = thin
            b = cell.border
            cell.border = Border(
                top    = _s(top)    if top    else b.top,
                bottom = _s(bottom) if bottom else b.bottom,
                left   = _s(left)   if left   else b.left,
                right  = _s(right)  if right  else b.right,
            )

def build_excel(entries_by_date: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    col_widths = [14, 8, 8, 10, 13, 4, 11, 4, 8, 8, 10]
    for date_key in sorted(entries_by_date.keys()):
        entries = entries_by_date[date_key]
        try:
            d = datetime.datetime.strptime(date_key, '%Y-%m-%d')
            sheet_name = d.strftime('%d-%b')
        except Exception:
            sheet_name = date_key
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.sheet_view.showGridLines = False
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        current_row = 1
        for entry in entries:
            write_entry(ws, current_row, entry)
            current_row += 7
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── Routes ────────────────────────────────────────────────────────────────────

HTML = open('index.html', encoding='utf-8').read()

@app.route('/')
def index():
    return Response(HTML, mimetype='text/html')

@app.route('/entrada', methods=['POST'])
def save_entrada():
    e = request.get_json()
    importe = float(e.get('importe') or 0)
    descuento = float(e.get('descuento') or 0)
    
    # Calculate commission based on discount percentage
    if 0 <= descuento <= 5:
        tasa_comision = 0.03
    elif 5 < descuento <= 10:
        tasa_comision = 0.025
    elif 10 < descuento <= 15:
        tasa_comision = 0.015
    else:
        tasa_comision = 0.0
        
    venta_total = importe * (1 - descuento / 100.0)
    comision = importe * tasa_comision

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO entradas
                  (fecha, nombre, razon_social, nombre_obra, telefono,
                   ubicacion, tipos, motivos, otro_detalle, m3_concreto, no_venta_block,
                   importe, descuento, comision)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
            """, (
                e.get('fecha'), e.get('nombre'), e.get('razonSocial'),
                e.get('nombreObra'), e.get('telefono'), e.get('ubicacion'),
                json.dumps(e.get('tipos', [])), json.dumps(e.get('motivos', [])),
                e.get('otroDetalle'), e.get('m3Concreto'), e.get('noVentaBlock'),
                importe, descuento, comision
            ))
            new_id = cur.fetchone()[0]
        conn.commit()
    return jsonify({'id': new_id}), 201

@app.route('/entradas')
def get_entradas():
    fecha_desde = request.args.get('desde')
    fecha_hasta = request.args.get('hasta')
    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            if fecha_desde and fecha_hasta:
                cur.execute("""
                    SELECT * FROM entradas
                    WHERE fecha BETWEEN %s AND %s
                    ORDER BY fecha, id
                """, (fecha_desde, fecha_hasta))
            else:
                cur.execute("SELECT * FROM entradas ORDER BY fecha DESC, id DESC LIMIT 200")
            rows = cur.fetchall()
    result = []
    for r in rows:
        result.append({
            'id': r['id'],
            'fecha': r['fecha'],
            'nombre': r['nombre'],
            'razonSocial': r['razon_social'] or '',
            'nombreObra': r['nombre_obra'] or '',
            'telefono': r['telefono'] or '',
            'ubicacion': r['ubicacion'] or '',
            'tipos': r['tipos'] if isinstance(r['tipos'], list) else json.loads(r['tipos'] or '[]'),
            'motivos': r['motivos'] if isinstance(r['motivos'], list) else json.loads(r['motivos'] or '[]'),
            'otroDetalle': r['otro_detalle'] or '',
            'm3Concreto': r['m3_concreto'] or '',
            'noVentaBlock': r['no_venta_block'] or '',
            'importe': float(r['importe'] or 0) if 'importe' in r and r['importe'] is not None else 0.0,
            'descuento': float(r['descuento'] or 0) if 'descuento' in r and r['descuento'] is not None else 0.0,
            'comision': float(r['comision'] or 0) if 'comision' in r and r['comision'] is not None else 0.0,
        })
    return jsonify(result)

@app.route('/entrada/<int:entry_id>', methods=['DELETE'])
def delete_entrada(entry_id):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM entradas WHERE id = %s", (entry_id,))
        conn.commit()
    return jsonify({'ok': True})

@app.route('/entrada/<int:entry_id>', methods=['PUT'])
def update_entrada_venta(entry_id):
    e = request.get_json()
    importe = float(e.get('importe') or 0)
    descuento = float(e.get('descuento') or 0)
    
    # Calculate commission based on discount percentage
    if 0 <= descuento <= 5:
        tasa_comision = 0.03
    elif 5 < descuento <= 10:
        tasa_comision = 0.025
    elif 10 < descuento <= 15:
        tasa_comision = 0.015
    else:
        tasa_comision = 0.0
        
    venta_total = importe * (1 - descuento / 100.0)
    comision = importe * tasa_comision

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE entradas
                SET importe = %s, descuento = %s, comision = %s
                WHERE id = %s
            """, (importe, descuento, comision, entry_id))
        conn.commit()
    return jsonify({'ok': True})

@app.route('/export', methods=['POST'])
def export():
    data = request.get_json()
    entries_by_date = {}
    for e in data.get('entries', []):
        entries_by_date.setdefault(e['fecha'], []).append(e)
    xlsx_bytes = build_excel(entries_by_date)
    return send_file(
        io.BytesIO(xlsx_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='reporte.xlsx'
    )

if __name__ == '__main__':
    app.run(debug=False)
